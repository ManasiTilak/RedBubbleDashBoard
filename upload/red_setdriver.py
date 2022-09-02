from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import glob
import os
import openpyxl

def start_reddriver(username, email, password, folder, excel):
    #setting webdriver and removing *Chrome is being controlled by automated test software* infobar
    options = webdriver.ChromeOptions()
    options.add_experimental_option("useAutomationExtension", False)
    options.add_experimental_option("excludeSwitches",["enable-automation"])

    #setting driver path
    driver_path = 'C:\Program Files (x86)\chromedriver.exe'
    driver = webdriver.Chrome(executable_path=driver_path, chrome_options=options)

    #setting driver path
    driver_path = 'C:\Program Files (x86)\chromedriver.exe'
    driver = webdriver.Chrome(executable_path=driver_path, chrome_options=options)

    #set driver
    driver.get("https://www.redbubble.com/auth/login")

    sign_in(username, password, folder, excel, driver)

def sign_in(username, password, folder, excel, driver):
    #sending username
    user_name = driver.find_element(By.ID, "ReduxFormInput1")
    user_name.send_keys(username)

    #sending password
    pass_word = driver.find_element(By.ID, "ReduxFormInput2")
    pass_word.send_keys(password)
    pass_word.send_keys(Keys.RETURN)
    time.sleep(15)
    # opening excel file
    workbook = openpyxl.load_workbook(excel)
    object = workbook['Sheet1']
    
    images = os.listdir(folder)
    for image in images:
        
        #getting image address
        file = folder + image
        
        driver.get("https://www.redbubble.com/portfolio/images/new")
        driver.find_element(By.CLASS_NAME, "copy-icon").click()

        goto_Button = driver.find_element(By.CLASS_NAME,"works_work-menu-link")
        goto_Button.click()
        time.sleep(10)
    #navigating the drop-down to Copy settings.
        try:
            optionsMenus = driver.find_elements(By.CLASS_NAME, "works_work-menu")
            for eachMenu in optionsMenus:
                    anchors = eachMenu.find_elements(By.TAG_NAME, "a")
                    anchors[2].click()
            time.sleep(10)

        except Exception as e:
            print(e)

        time.sleep(10)
        #filling the form
        # Title
        title = driver.find_element(By.ID, "work_title_en")
        title.clear()
        tit = object.cell(1, 1).value
        title.send_keys(tit)
        #Tags
        tagss = driver.find_element(By.ID, "work_tag_field_en")
        tagss.clear()
        tg = object.cell(1, 2).value
        tagss.send_keys(tg)
        #description
        desc = driver.find_element(By.ID, "work_description_en")
        desc.clear()
        des = object.cell(1, 3).value
        desc.send_keys(des)
        #image
        img = driver.find_element(By.ID, "select-image-base")
        img.send_keys(file)
        driver.find_element(By.ID, "rightsDeclaration").click()
        time.sleep(50)
        driver.find_element(By.ID, "submit-work").click()
        #delete the first row. The second row becomes the first row
        object.delete_rows(1)
        #setting wait until to start new upload
        urlpattern = 'https://www.redbubble.com/studio/promote'
        wait = WebDriverWait(driver, 50)
        wait.until(EC.url_contains(urlpattern))
    #save the excel file
    workbook.save(excel)
    
    time.sleep(10)

    driver.quit()