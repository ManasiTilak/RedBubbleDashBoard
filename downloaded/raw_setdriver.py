def start_driver(range_upper):
    import os
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    import time
    from openpyxl import load_workbook
    from dotenv import load_dotenv, find_dotenv

    #loading env file
    load_dotenv(find_dotenv())

    #set driver
    driver_path = 'C:\Program Files (x86)\chromedriver.exe'
    driver = webdriver.Chrome(executable_path=driver_path)
    #set sheet
    WBR = load_workbook(r'.\res\rbdash_links.xlsx')
    OBJR = WBR['Sheet1']
    
    EMAIL_ADDRESS = os.getenv('EMAIL_ADDRESS')
    PASSWORD = os.getenv('PASSWORD') 

    driver.get("https://www.rawpixel.com/user/login")
    

    #sending username
    driver.find_element(By.XPATH, "/html/body/div[1]/div/main/div[2]/div[1]/form/div[1]/input").send_keys(EMAIL_ADDRESS)
    #sending password
    password = driver.find_element(By.ID, "edit-pass")
    password.send_keys(PASSWORD)
    password.send_keys(Keys.RETURN)
    #pressing enter
    #navigate to the public domain side
    time.sleep(30)
    #starting download
    for i in range(1,range_upper+1):
        #get the current cell
        print(i)
        current_cell = OBJR.cell(1,1)
        link = current_cell.value
        print(link)
        driver.get(link)
        time.sleep(30)
        downloadbtn = driver.find_element(By.XPATH, "/html/body/div[1]/div/main/div[2]/article/figure/figcaption/div/div[1]/div/button/span")
        downloadbtn.click()
        time.sleep(15)
        # deleting the specific link
        OBJR.delete_rows(1)
        WBR.save(r'.\res\rbdash_links.xlsx')
    time.sleep(50)
    driver.quit()
    
    print("Images downloaded. Please Check")
