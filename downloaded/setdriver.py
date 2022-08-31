def start_driver(range_upper):
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    import time
    from openpyxl import load_workbook

    #set driver
    driver_path = 'C:\Program Files (x86)\chromedriver.exe'
    driver = webdriver.Chrome(executable_path=driver_path)
    #set sheet
    WBR = load_workbook(r'.\res\rbdash_links.xlsx')
    OBJR = WBR['Sheet1']
    driver.get("https://www.rawpixel.com/user/login")
    

    #sending username
    driver.find_element(By.XPATH, "/html/body/div[1]/div/main/div[2]/div[1]/form/div[1]/input").send_keys("manasitilak12@gmail.com")
    #sending password
    password = driver.find_element(By.ID, "edit-pass")
    password.send_keys("Sana2020")
    password.send_keys(Keys.RETURN)
    #pressing enter
    #navigate to the public domain side
    time.sleep(30)
    #starting download
    for i in range(1,range_upper+1):
        #get the current cell
        print(i)
        current_cell = OBJR.cell(1,1)
        link = current_cell.value
        print(link)
        driver.get(link)
        time.sleep(30)
        downloadbtn = driver.find_element(By.CLASS_NAME, "sc-2954edd3-2")
        downloadbtn.click()
        time.sleep(15)
        # deleting the specific link
        OBJR.delete_rows(1)
    time.sleep(50)
    driver.quit()
    WBR.save(r'.\res\try_links.xlsx')
    print("Images downloaded. Please Check")
