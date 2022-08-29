def transfer_links(upper_range):
    from openpyxl import load_workbook
    WB = load_workbook(r'./res/trylinks.xlsx')
    OBJ = WB['Sheet1']

    with open("scrape_links_modified.txt", "r+") as f:
        link = f.readlines()
        f.truncate(0)

    for i in range(0, upper_range): #upper limit is the number of lines in the modified text file
        cell_position = OBJ.cell(row = i+1, column = 1)
        link_text = link[i]
        cell_position.value = link_text

    print("Links have been transferred. You may proceed with downloading images. All the best")
    WB.save(r'./res/trylinks.xlsx')